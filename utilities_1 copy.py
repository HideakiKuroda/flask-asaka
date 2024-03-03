from flask import Flask, request, request,flash, redirect, url_for
from openpyxl import load_workbook
from flask import session
import os
import win32com.client as win32
import pythoncom
import logging
from datetime import datetime, timedelta,time

#app.py で form_dataに格納されたデータをdata としてExcelファイルに書き込む
def edit_excel(data):
    exfilename = session.get('filename')
    if not exfilename:
        return 'Filename is not set in session.'
    exfilename = os.path.join('dailyWorkReports', exfilename)
    book = load_workbook(exfilename)
    sheet = book.active
    # セルにデータを挿入
    sheet['F4'] = data['category']
    date_obj = datetime.strptime(data['date'], '%Y-%m-%d')
    formatted_date = date_obj.strftime('%Y年%m月%d日')
    sheet['B4'] = formatted_date
    sheet['C4'] = data['weekday']
    sheet['Q4'] = data['person']
    sheet['C22'] = data['opening']
    sheet['C23'] = data['closed']

    # 船舶詳細データの書き込み
    row = 8  # 開始行
    for work_detail in data['work_details']:
        if row > 20:
            break  # 20行を超えたらループを終了
        sheet[f'A{row}'] = work_detail.get('shipname')
        sheet[f'E{row}'] = work_detail.get('berth')
        sheet[f'F{row}'] = work_detail.get('details')
        sheet[f'G{row}'] = work_detail.get('schedule')
        sheet[f'H{row}'] = work_detail.get('departure')
        sheet[f'I{row}'] = work_detail.get('onsite')
        sheet[f'J{row}'] = work_detail.get('start')
        sheet[f'K{row}'] = work_detail.get('end')
        sheet[f'L{row}'] = work_detail.get('arrival')
        sheet[f'M{row}'] = work_detail.get('usage')
        sheet[f'N{row}'] = work_detail.get('certificate')
        sheet[f'N{row-1}'] = work_detail.get('partner')
        row += 2  # 次の入力行を設定（一行飛ばし）

    book.save(exfilename)
    
    return f'Excel ファイル {exfilename} に書き込みが完了しました.'

#Excelファイルに書き込まれたデータを読み込む
def intake_from_exl(filename):
    exfilename = filename
    if not exfilename:
        return 'Filename is not set in session.'
    exfilename = os.path.join('dailyWorkReports', exfilename)
    book = load_workbook(exfilename)
    sheet = book.active

    excel_data = {
    'date' : sheet['B4'].value,
    'weekday' : sheet['C4'].value,
    'category' : sheet['F4'].value,
    'person' : sheet['Q4'].value,
    'opening': sheet['C22'].value,
    'closed' : sheet['C23'].value,
    'work_details' :[]
    }
    row = 8
    while row <= 20:
        shipname = sheet[f'A{row}'].value
        if not shipname:  # 船名がなければループを終了
            break
        work_detail = {
            'shipname': shipname,
            'berth': sheet[f'E{row}'].value,
            'details': sheet[f'F{row}'].value,
            'schedule': sheet[f'G{row}'].value,
            'departure': sheet[f'H{row}'].value,
            'onsite': sheet[f'I{row}'].value,
            'start': sheet[f'J{row}'].value,
            'end': sheet[f'K{row}'].value,
            'arrival': sheet[f'L{row}'].value,
            'usage': sheet[f'M{row}'].value,
            'certificate': sheet[f'N{row}'].value,
            'partner': sheet[f'N{row-1}'].value  # パートナー情報は1行下
        }
        excel_data['work_details'].append(work_detail)
        row += 2  # 次の船舶情報へ（一行飛ばし）
        book.save(exfilename)
    return excel_data

def generate_new_filename(base_path):
    # ファイルの基本名と拡張子を分離
    base, extension = os.path.splitext(base_path)
    counter = 1  # 連番の開始

    # 新しいファイル名を生成
    new_file_path = f"{base}({counter}){extension}"

    # 生成したファイル名が既に存在する場合は、連番を増やして再試行
    while os.path.exists(new_file_path):
        flash('ファイルが既に存在します！') 
        counter += 1
        new_file_path = f"{base}({counter}){extension}"

    return new_file_path

def print_excel_file(file_path):
    # Excelアプリケーションを開始
    pythoncom.CoInitialize()   #特にマルチスレッド環境でそのスレッドでCOMコンポーネントが適切に初期化
    try:
        # Excelアプリケーションを開始
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        #相対パスを絶対パスに変換（Excelでファイルを読み込む場合は必要）
        absolute_path = os.path.abspath(file_path)
        # ファイルを開く（Excelアプリケーションは非表示）
        excel.Visible = False
        workbook = excel.Workbooks.Open(absolute_path)
        
        try:
            # アクティブなシートを印刷（既定のプリンターを使用）
            excel.ActiveSheet.PrintOut()
        finally:
            # ワークブックを閉じる（変更を保存せず）
            workbook.Close(SaveChanges=False)
    finally:
        # Excelアプリケーションを終了
        excel.Quit()
        # COMライブラリの使用を終了（重要）
    pythoncom.CoUninitialize()
        
def print_totalling_file(file_path):
    # Excelアプリケーションを開始
    pythoncom.CoInitialize()   #特にマルチスレッド環境でそのスレッドでCOMコンポーネントが適切に初期化
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        absolute_path = os.path.abspath(file_path)
        excel.Visible = False
        workbook = excel.Workbooks.Open(absolute_path)

        # QRS列を非表示に設定
        for col in ['Q', 'R', 'S']:
            excel.Columns(col).Hidden = True

        # 印刷範囲をA1～V24に変更
        excel.ActiveSheet.PageSetup.PrintArea = "$A$2:$V$24"
        
        try:
            # アクティブなシートを印刷（既定のプリンターを使用）
            excel.ActiveSheet.PrintOut()
        finally:
            # ワークブックを閉じる（変更を保存せず）
            workbook.Close(SaveChanges=False)
    finally:
        # Excelアプリケーションを終了
        excel.Quit()
        # COMライブラリの使用を終了（重要）
    pythoncom.CoUninitialize()
            
# 日勤の場合の時間外計算
def dayshift_overtime_to_excel(file_path):
    # Excelファイルを開く
    book = load_workbook(file_path)
    sheet = book.active
    # 就業時間帯の定義
    work_periods = [
        (datetime.strptime('06:00', '%H:%M'), datetime.strptime('10:00', '%H:%M')),
        (datetime.strptime('15:00', '%H:%M'), datetime.strptime('19:00', '%H:%M'))
    ]
    # 対象の行リスト
    rows = [8, 10, 12, 14, 16, 18, 20]
    tt_overtime  = timedelta()
    for row in rows:
        start = sheet[f'H{row}'].value
        end = sheet[f'L{row}'].value
        if not start or not end:
            continue  # 空のセルはスキップ
        start_time = datetime.strptime(start, '%H:%M')
        end_time = datetime.strptime(end, '%H:%M')
        overtime = timedelta()
        # 作業時間が就業時間帯の間に完全にある場合、直接時間外労働として計算
        in_between = True
        for work_start, work_end in work_periods:
            if start_time < work_end and end_time > work_start:
                in_between = False
                # 作業開始が就業開始前、または作業終了が就業終了後の場合
                if start_time < work_start:
                    overtime += work_start - start_time
                if end_time > work_end:
                    if end_time > datetime.strptime('15:00', '%H:%M') and work_end == datetime.strptime('10:00', '%H:%M'):
                        overtime += timedelta(hours=5) 
                    else: overtime += end_time - work_end
                start_time = max(start_time, work_end)
            elif start_time >= work_end:
                continue
            elif end_time <= work_start:
                break
        if in_between:
            overtime += end_time - start_time
        tt_overtime += overtime
        overtime_minutes = overtime.total_seconds() / 60
        overtime_formatted = minutes_to_hours_and_minutes(overtime_minutes)
        sheet[f'T{row}'].value = overtime_formatted
    tt_overtime_minutes = tt_overtime.total_seconds() / 60
    tt= minutes_to_hours_and_minutes(tt_overtime_minutes)
    sheet['V8'].value = tt
    #日勤時間の午前午後に4:00 4:00 を入力
    sheet['V6'].value = datetime.strptime('04:00', '%H:%M').time()  
    sheet['V7'].value = datetime.strptime('04:00', '%H:%M').time()
      
    book.save(file_path)

#当直の場合の時間外計算
def onduty_overtime_to_excel(file_path):
	# Excelファイルを開く
    book = load_workbook(file_path)
    sheet = book.active
	# 就業時間帯の定義
    work_periods = [
        (datetime.strptime('06:00', '%H:%M'), datetime.strptime('10:00', '%H:%M')),
        (datetime.strptime('15:00', '%H:%M'), datetime.strptime('23:59', '%H:%M') + timedelta(days=1))
    ]
	# 対象の行リスト
    rows = [8, 10, 12, 14, 16, 18, 20]
    tt_overtime  = timedelta()
    for row in rows:
        start = sheet[f'H{row}'].value
        end = sheet[f'L{row}'].value
        if not start or not end:
            continue  # 空のセルはスキップ
        start_time = datetime.strptime(start, '%H:%M')
        end_time = datetime.strptime(end, '%H:%M')
        overtime = timedelta()
		# 作業時間が就業時間帯の間に完全にある場合、直接時間外労働として計算
        in_between = True
        for work_start, work_end in work_periods:
            if start_time < work_end and end_time > work_start:#====================================>>>>>>>ここを修正！！！！！
                in_between = False
				# 作業開始が就業開始前、または作業終了が就業終了後の場合
                if start_time < work_start:
                    overtime += work_start - start_time
                if end_time > work_end:
                    if end_time > datetime.strptime('15:00', '%H:%M') and work_end == datetime.strptime('10:00', '%H:%M'):
                        overtime += timedelta(hours=5) 
                    else: overtime += end_time - work_end
                start_time = max(start_time, work_end)
            elif start_time >= work_end:
                continue
            elif end_time <= work_start:
                break
        if in_between:
            overtime += end_time - start_time
        tt_overtime += overtime
        overtime_minutes = overtime.total_seconds() / 60
        overtime_formatted = minutes_to_hours_and_minutes(overtime_minutes)
        sheet[f'T{row}'].value = overtime_formatted
    tt_overtime_minutes = tt_overtime.total_seconds() / 60
    tt= minutes_to_hours_and_minutes(tt_overtime_minutes)
    sheet['V8'].value = tt 
	#日勤時間の午前午後に4:00 4:00 を入力
    sheet['V6'].value = time(hour=4, minute=0)  
    sheet['V7'].value = time(hour=4, minute=0)

	# C23が空欄の場合、V9に3:00、V10に2:00を設定
    if sheet['C23'].value is None:
        sheet['V9'].value = time(hour=3, minute=0)  # 3:00
        sheet['V10'].value = time(hour=2, minute=0)  # 2:00
	# 変更を保存
    book.save(file_path)

#当直明けの時間外計算    
def endofshift_overtime_to_excel(file_path):
    # Excelファイルを開く
	book = load_workbook(file_path)
	sheet = book.active
	closed_time = datetime.strptime(request.form['closed'], '%H:%M')
	base_time = datetime.strptime('5:00', '%H:%M')
	if closed_time < base_time:
        # '5:00'から'closed'の差を計算し、V11に設定
		delta = base_time - closed_time
		minutes = delta.seconds / 60
		sheet['V11'].value = minutes_to_hours_and_minutes(minutes)
		sheet['V12'].value = time(hour=0, minute=0)  #'0:00' V12には0時間を設定
	else:
        # 'closed'が'5:00'以降の場合
		sheet['V11'].value = time(hour=5, minute=0) # '5:00' V11には5時間を設定
		if closed_time > base_time:
            # 'closed'から'5:00'の差を計算し、V12に設定
			delta = closed_time - base_time
			minutes = delta.seconds / 60
			sheet['V12'].value = minutes_to_hours_and_minutes(minutes)
		else:
			sheet['V12'].value = time(hour=0, minute=0)  #'0:00'  同じ場合は0時間
	book.save(file_path)

# 分を時間と分に変換する関数
def minutes_to_hours_and_minutes(minutes):
    hours, minutes = divmod(minutes, 60)
    return time(hour=int(hours), minute=int(minutes))

logging.basicConfig(filename='app.log', level=logging.INFO, 
                    format='%(asctime)s %(levelname)s:%(message)s')
# 実際にファイルを更新する場合はこの関数を呼び出します（ファイルパスを適切に設定してください）
# update_overtime_in_excel('your_excel_file.xlsx')

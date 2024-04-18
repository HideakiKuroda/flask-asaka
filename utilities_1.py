from flask import Flask, request, request,flash, redirect, url_for
from openpyxl import load_workbook
from flask import session
import os
import win32com.client as win32
import pythoncom
import logging
from datetime import datetime, timedelta,time 

#app.py で form_dataに格納されたデータをdata としてExcelファイルに書き込む
def edit_excel(data):
    exfilename = session.get('filename')
    if not exfilename:
        return 'Filename is not set in session.'
    exfilename = os.path.join('dailyWorkReports', exfilename)
    book = load_workbook(exfilename)
    sheet = book.active
    # セルにデータを挿入
    sheet['F4'] = data['category']
    date_obj = datetime.strptime(data['date'], '%Y-%m-%d')
    formatted_date = date_obj.strftime('%Y年%m月%d日')
    sheet['B4'] = formatted_date
    sheet['C4'] = data['weekday']
    sheet['Q4'] = data['person']
    sheet['C22'] = data['opening']
    sheet['C23'] = data['closed']
    sheet['F22'] = data.get('remarks1', '')  # 'remarks1'がない場合は空文字を設定
    sheet['F23'] = data.get('remarks2', '')  # 'remarks2'がない場合は空文字を設定

    # 船舶詳細データの書き込み
    row = 8  # 開始行
    for work_detail in data['work_details']:
        if row > 20:
            break  # 20行を超えたらループを終了
        sheet[f'A{row}'] = work_detail.get('shipname')
        sheet[f'E{row}'] = work_detail.get('berth')
        sheet[f'F{row}'] = work_detail.get('details')
        sheet[f'G{row}'] = work_detail.get('schedule')
        sheet[f'H{row}'] = work_detail.get('departure')
        sheet[f'I{row}'] = work_detail.get('onsite')
        sheet[f'J{row}'] = work_detail.get('start')
        sheet[f'K{row}'] = work_detail.get('end')
        sheet[f'L{row}'] = work_detail.get('arrival')
        sheet[f'M{row}'] = work_detail.get('usage')
        sheet[f'N{row}'] = work_detail.get('certificate')
        sheet[f'N{row-1}'] = work_detail.get('partner')
        row += 2  # 次の入力行を設定（一行飛ばし）

    book.save(exfilename)
    
    return f'Excel ファイル {exfilename} に書き込みが完了しました.'

# 分を時間と分に変換する関数
def minutes_to_hours_and_minutes(minutes):
    hours, minutes = divmod(minutes, 60)
    return time(hour=int(hours), minute=int(minutes))

#Excelファイルに書き込まれたデータを読み込む
def intake_from_exl(filename):
    exfilename = filename
    if not exfilename:
        return 'Filename is not set in session.'
    exfilename = os.path.join('dailyWorkReports', exfilename)
    book = load_workbook(exfilename)
    sheet = book.active

    excel_data = {
    'date' : sheet['B4'].value,
    'weekday' : sheet['C4'].value,
    'category' : sheet['F4'].value,
    'person' : sheet['Q4'].value,
    'opening': sheet['C22'].value,
    'closed' : sheet['C23'].value,
    'remarks1' : sheet['F22'].value if sheet['F22'].value is not None else '',
    'remarks2' : sheet['F23'].value if sheet['F23'].value is not None else '',
    'work_details' :[]
    }
    row = 8
    while row <= 20:
        shipname = sheet[f'A{row}'].value
        if not shipname:  # 船名がなければループを終了
            break
        work_detail = {
            'shipname': shipname,
            'berth': sheet[f'E{row}'].value,
            'details': sheet[f'F{row}'].value,
            'schedule': sheet[f'G{row}'].value,
            'departure': sheet[f'H{row}'].value,
            'onsite': sheet[f'I{row}'].value,
            'start': sheet[f'J{row}'].value,
            'end': sheet[f'K{row}'].value,
            'arrival': sheet[f'L{row}'].value,
            'usage': sheet[f'M{row}'].value,
            'certificate': sheet[f'N{row}'].value,
            'partner': sheet[f'N{row-1}'].value  # パートナー情報は1行下
        }
        excel_data['work_details'].append(work_detail)
        row += 2  # 次の船舶情報へ（一行飛ばし）
        book.save(exfilename)
    return excel_data

def generate_new_filename(base_path):
    # ファイルの基本名と拡張子を分離
    base, extension = os.path.splitext(base_path)
    counter = 1  # 連番の開始

    # 新しいファイル名を生成
    new_file_path = f"{base}({counter}){extension}"

    # 生成したファイル名が既に存在する場合は、連番を増やして再試行
    while os.path.exists(new_file_path):
        flash('ファイルが既に存在します！') 
        counter += 1
        new_file_path = f"{base}({counter}){extension}"

    return new_file_path

def print_excel_file(file_path):
    # Excelアプリケーションを開始
    pythoncom.CoInitialize()   #特にマルチスレッド環境でそのスレッドでCOMコンポーネントが適切に初期化
    try:
        # Excelアプリケーションを開始
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        #相対パスを絶対パスに変換（Excelでファイルを読み込む場合は必要）
        absolute_path = os.path.abspath(file_path)
        # ファイルを開く（Excelアプリケーションは非表示）
        excel.Visible = False
        workbook = excel.Workbooks.Open(absolute_path)
        
        try:
            # アクティブなシートを印刷（既定のプリンターを使用）
            excel.ActiveSheet.PrintOut()
        finally:
            # ワークブックを閉じる（変更を保存せず）
            workbook.Close(SaveChanges=False)
    finally:
        # Excelアプリケーションを終了
        excel.Quit()
        # COMライブラリの使用を終了（重要）
    pythoncom.CoUninitialize()
        
def print_totalling_file(file_path):
    # Excelアプリケーションを開始
    pythoncom.CoInitialize()   #特にマルチスレッド環境でそのスレッドでCOMコンポーネントが適切に初期化
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        absolute_path = os.path.abspath(file_path)
        excel.Visible = False
        workbook = excel.Workbooks.Open(absolute_path)

        # QRS列を非表示に設定
        for col in ['Q', 'R', 'S']:
            excel.Columns(col).Hidden = True

        # 印刷範囲をA1～V24に変更
        excel.ActiveSheet.PageSetup.PrintArea = "$A$2:$V$24"
        
        try:
            # アクティブなシートを印刷（既定のプリンターを使用）
            excel.ActiveSheet.PrintOut()
        finally:
            # ワークブックを閉じる（変更を保存せず）
            workbook.Close(SaveChanges=False)
    finally:
        # Excelアプリケーションを終了
        excel.Quit()
        # COMライブラリの使用を終了（重要）
    pythoncom.CoUninitialize()
            
# 日勤の場合の時間外計算
def dayshift_overtime_to_excel(file_path):
    # Excelファイルを開く
    book = load_workbook(file_path)
    sheet = book.active
    # 就業時間帯の定義
    work_periods = [
        (datetime.strptime('06:00', '%H:%M'), datetime.strptime('10:00', '%H:%M')),
        (datetime.strptime('15:00', '%H:%M'), datetime.strptime('19:00', '%H:%M'))
    ]
    # 対象の行リスト
    rows = [8, 10, 12, 14, 16, 18, 20]
    tt_overtime  = timedelta()
    tt_midtime  = timedelta()
    tt_earlytime = timedelta()
    tt_usetime = timedelta()
    for row in rows:
        usem = sheet[f'M{row}'].value
        if usem:  # usemがNoneまたは空文字列でないことを確認
            formatted_time = format_time_to_str(usem)
            # 時間をtimedeltaオブジェクトに変換
            hours, minutes = map(int, formatted_time.split(':'))
            usetime = timedelta(hours=hours, minutes=minutes)
            # tt_usetimeに加算
            tt_usetime += usetime
        #前の作業の終了時間を取得、最初の行の時は’’
        start = sheet[f'H{row}'].value
        end = sheet[f'L{row}'].value
        if not start or not end:
            continue  # 空のセルはスキップ
        start_time = datetime.strptime(format_time_to_str(start), '%H:%M') #format_time_to_str(
        end_time = datetime.strptime(format_time_to_str(end), '%H:%M')
        if row >= 10:
            pre_end = sheet[f'L{row - 2}'].value
            if pre_end:  # pre_endが空でない場合
                pre_end_time_str = format_time_to_str(pre_end)
                pre_end_time = get_datetime_object(pre_end_time_str)
                # start_timeが未定義の場合、比較はスキップされるべきか、
                # または適切な初期値を設定する必要がある
                if start_time.time() < pre_end_time.time():
                    start_time = pre_end_time  # 作業開始時間を更新
        early_time = datetime.strptime('05:00', '%H:%M')
        mid_time = datetime.strptime('22:00', '%H:%M')
        overtime = timedelta()
        early_time_work = timedelta()
        mid_time_work = timedelta()
        # 作業時間が就業時間帯の間に完全にある場合、直接時間外労働として計算
        in_between = True
        #5:00より前に作業があった場合
        if start_time < early_time:
            early_time_work += early_time - start_time
        #22:00より後に作業があった場合    
        if end_time > mid_time:
            mid_time_work += end_time - mid_time    
        # 勤務時間が両方の就業時間帯をまたいでいる場合の特別な処理
        if start_time < work_periods[0][1] and end_time > work_periods[1][0]:
            # 朝の就業時間帯の終了後から夜の就業時間帯の開始前までを時間外労働として計算
            overtime += work_periods[1][0] - work_periods[0][1]
            # 勤務開始時間が最初の就業時間帯の開始時間よりも前である場合の時間外労働を計算
            if start_time < work_periods[0][0]:
                overtime += work_periods[0][0] - start_time
            # 勤務終了時間が最後の就業時間帯の終了時間よりも後である場合の時間外労働を計算
            if end_time > work_periods[-1][1]:
                overtime += end_time - work_periods[-1][1]        
        else: #勤務時間が両方の就業時間帯をまたいでいない通常の処理
            for work_start, work_end in work_periods:
                if start_time < work_end and end_time > work_start: #勤務時間帯が定時の時間に掛かっている場合
                    in_between = False
                    if start_time < work_start:
                        overtime += work_start - start_time
                    if end_time > work_end:
                        overtime += end_time - work_end
                elif end_time <= work_start:
                    break
            if in_between: #勤務時間帯が定時の時間に掛かっていない場合はすべてが時間外
                overtime += end_time - start_time
        #記入する時間外の時間は深夜早朝も含める
        overtime_minutes = overtime.total_seconds() / 60
        overtime_formatted = minutes_to_hours_and_minutes(overtime_minutes)
        sheet[f'T{row}'].value = overtime_formatted
        #全ての時間外から深夜と早朝の時間外を引くことで普通時間外の合計
        tt_overtime += overtime -  early_time_work - mid_time_work
        #深夜の時間外合計
        tt_midtime += mid_time_work
        #早朝の時間外の合計
        tt_earlytime += early_time_work
    #普通時間外の合計
    tt_overtime_minutes = tt_overtime.total_seconds() / 60
    tt= minutes_to_hours_and_minutes(tt_overtime_minutes)
    sheet['V8'].value = tt
    #深夜時間外の合計
    tt_midtime_minutes = tt_midtime.total_seconds() / 60
    tt_mid= minutes_to_hours_and_minutes(tt_midtime_minutes) 
    sheet['V10'].value = tt_mid
    #早朝時間外の合計
    tt_earlytime_minutes = tt_earlytime.total_seconds() / 60
    tt_early = minutes_to_hours_and_minutes(tt_earlytime_minutes) 
    sheet['V11'].value = tt_early
    #日勤時間の午前午後に4:00 4:00 を入力
    sheet['V6'].value = time(hour=4, minute=0)   
    sheet['V7'].value = time(hour=4, minute=0)  
    sheet['V13'].value = minutes_to_hours_and_minutes(tt_overtime_minutes + tt_midtime_minutes +  tt_earlytime_minutes +240 + 240)  #合計時間（全て）  
    sheet['V15'].value = minutes_to_hours_and_minutes(tt_overtime_minutes) #普通の時間外合計時間
    sheet['V16'].value =minutes_to_hours_and_minutes(tt_midtime_minutes +  tt_earlytime_minutes) #深夜の時間外合計時間
    tt_usetime_sec = tt_usetime.total_seconds() / 60
    sheet['V19'].value = minutes_to_hours_and_minutes(tt_usetime_sec) #普通の時間外合計時間
    book.save(file_path)

#当直の場合の時間外計算
def onduty_overtime_to_excel(file_path):
	# Excelファイルを開く
    book = load_workbook(file_path)
    sheet = book.active
	# 就業時間帯の定義
    work_periods = [
        (datetime.strptime('06:00', '%H:%M'), datetime.strptime('10:00', '%H:%M')),
        (datetime.strptime('15:00', '%H:%M'), datetime.strptime('23:59', '%H:%M'))
    ]
	# 対象の行リスト
    rows = [8, 10, 12, 14, 16, 18, 20]
    tt_overtime  = timedelta()
    tt_earlytime = timedelta()
    tt_usetime = timedelta()
    midnight = datetime.strptime('00:00', '%H:%M')
    for row in rows:
        usem = sheet[f'M{row}'].value
        if usem:  # usemがNoneまたは空文字列でないことを確認
            formatted_time = format_time_to_str(usem)
            # 時間をtimedeltaオブジェクトに変換
            hours, minutes = map(int, formatted_time.split(':'))
            usetime = timedelta(hours=hours, minutes=minutes)
            # tt_usetimeに加算
            tt_usetime += usetime
        #前の作業の終了時間を取得、最初の行の時は’’
        start = sheet[f'H{row}'].value
        end = sheet[f'L{row}'].value
        if not start or not end:
            continue  # 空のセルはスキップ
        start_time = datetime.strptime(format_time_to_str(start), '%H:%M') #format_time_to_str(
        end_time = datetime.strptime(format_time_to_str(end), '%H:%M')
        time_since_midnight = timedelta()
        if end_time < start_time:
            end_time += timedelta(hours=24)
            time_since_midnight = end_time - datetime.strptime('23:59', '%H:%M')
        if row >= 10:
            pre_end = sheet[f'L{row - 2}'].value
            if pre_end:  # pre_endが空でない場合
                pre_end_time_str = format_time_to_str(pre_end)
                pre_end_time = get_datetime_object(pre_end_time_str)
                # start_timeが未定義の場合、比較はスキップされるべきか、
                # または適切な初期値を設定する必要がある
                if start_time.time() < pre_end_time.time():
                    start_time = pre_end_time  # 作業開始時間を更新
        early_time = datetime.strptime('05:00', '%H:%M')
        overtime = timedelta()
        early_time_work = timedelta()
		# 作業時間が就業時間帯の間に完全にある場合、直接時間外労働として計算
        in_between = True
        #5:00より前に作業があった場合
        if start_time < early_time:
            early_time_work += early_time - start_time
        # 勤務時間が両方の就業時間帯をまたいでいる場合の特別な処理
        if start_time < work_periods[0][1] and end_time > work_periods[1][0]:
            # 朝の就業時間帯の終了後から夜の就業時間帯の開始前までを時間外労働として計算
            overtime += work_periods[1][0] - work_periods[0][1]
            # 勤務開始時間が最初の就業時間帯の開始時間よりも前である場合の時間外労働を計算
            if start_time < work_periods[0][0]:
                overtime += work_periods[0][0] - start_time
            # 勤務終了時間が最後の就業時間帯の終了時間よりも後である場合の時間外労働を計算
            if end_time > work_periods[-1][1]:
                overtime += end_time - work_periods[-1][1]        
        else: #勤務時間が両方の就業時間帯をまたいでいない通常の処理
            for work_start, work_end in work_periods:
                if start_time < work_end and end_time > work_start: #勤務時間帯が定時の時間に掛かっている場合
                    in_between = False
                    if start_time < work_start:
                        overtime += work_start - start_time
                    if end_time > work_end:
                        overtime += end_time - work_end
                elif end_time <= work_start:
                    break
            if in_between: #勤務時間帯が定時の時間に掛かっていない場合はすべてが時間外
                overtime += end_time - start_time
        if time_since_midnight:
            overtime -= time_since_midnight        
        overtime_minutes = overtime.total_seconds() / 60    
        overtime_formatted = minutes_to_hours_and_minutes(overtime_minutes)
        sheet[f'T{row}'].value = overtime_formatted
        #全ての時間外から深夜と早朝の時間外を引くことで普通時間外の合計
        tt_overtime += overtime - early_time_work
        #早朝の時間外の合計
        tt_earlytime += early_time_work
    #普通時間外の合計
    tt_overtime_minutes = tt_overtime.total_seconds() / 60
    tt= minutes_to_hours_and_minutes(tt_overtime_minutes)
    sheet['V8'].value = tt 
    #早朝時間外の合計
    tt_earlytime_minutes = tt_earlytime.total_seconds() / 60
    tt_early = minutes_to_hours_and_minutes(tt_earlytime_minutes) 
    sheet['V11'].value = tt_early

	#日勤時間の午前午後に4:00 4:00 を入力
    sheet['V6'].value = time(hour=4, minute=0)  
    sheet['V7'].value = time(hour=4, minute=0)
    sheet['V13'].value = minutes_to_hours_and_minutes(tt_overtime_minutes + tt_earlytime_minutes + 300 +240 + 240)  #合計時間（全て）  
    sheet['V15'].value = minutes_to_hours_and_minutes(tt_overtime_minutes) #普通の時間外合計時間
    sheet['V16'].value =minutes_to_hours_and_minutes(tt_earlytime_minutes) #早朝の時間外合計時間
    tt_usetime_sec = tt_usetime.total_seconds() / 60
    sheet['V19'].value = minutes_to_hours_and_minutes(tt_usetime_sec) #普通の時間外合計時間

	# C23が空欄の場合、V9に3:00、V10に2:00を設定
    if sheet['C23'].value is None:
        sheet['V9'].value = time(hour=3, minute=0)  # 3:00
        sheet['V10'].value = time(hour=2, minute=0)  # 2:00
        sheet['V15'].value = minutes_to_hours_and_minutes(tt_overtime_minutes + 180) #普通の時間外合計時間
        sheet['V16'].value = minutes_to_hours_and_minutes(120 + tt_earlytime_minutes) 
	# 変更を保存
    book.save(file_path)

#当直明けの時間外計算   閉局時間までが残業（深夜残業と普通残業） 
def endofshift_overtime_to_excel(file_path):
    # Excelファイルを開く
    book = load_workbook(file_path)
    sheet = book.active
    closed_time = datetime.strptime(format_time_to_str(request.form['closed']), '%H:%M')
    mid_time = datetime.strptime('00:00', '%H:%M')
    base_time = datetime.strptime('5:00', '%H:%M')
    mid_work = timedelta()
    dawn_work = timedelta()
    tt_dawn_time = timedelta()
    tt_usetime = timedelta()
    rows = [8, 10, 12, 14, 16, 18, 20]
    for row in rows:
        usem = sheet[f'M{row}'].value
        if usem:  # usemがNoneまたは空文字列でないことを確認
            formatted_time = format_time_to_str(usem)
            # 時間をtimedeltaオブジェクトに変換
            hours, minutes = map(int, formatted_time.split(':'))
            usetime = timedelta(hours=hours, minutes=minutes)
            # tt_usetimeに加算
            tt_usetime += usetime
    if closed_time < base_time: #閉局時間が5:00より早い場合
        # '5:00'から'closed'の差を計算し、V11に設定
        delta = closed_time - mid_time #'closed'までが深夜残業
        minutes = delta.seconds / 60
        sheet['V11'].value = minutes_to_hours_and_minutes(minutes) #深夜残業の時間
        sheet['V16'].value = minutes_to_hours_and_minutes(minutes) #深夜残業の時間合計
        mid_work = delta
        sheet['V12'].value = time(hour=0, minute=0)  #'0:00' V12には0時間を設定 5:00以降は作業なし
        sheet['V15'].value = time(hour=0, minute=0)  #'0:00' V12には0時間を設定 5:00以降は作業なし
        # dawn_work = time(hour=0, minute=0)
        dawn_work = timedelta()
    else: # 'closed'が'5:00'以降の場合
        sheet['V11'].value = time(hour=5, minute=0) # '5:00' V11には5時間を設定
        sheet['V16'].value = time(hour=5, minute=0) # '5:00' V11には5時間を設定
        mid_work = timedelta(hours=5) 
        if closed_time > base_time:
            # 'closed'から'5:00'の差を計算し、V12に設定
            delta = closed_time - base_time
            minutes = delta.seconds / 60
            sheet['V12'].value = minutes_to_hours_and_minutes(minutes)
            sheet['V15'].value = minutes_to_hours_and_minutes(minutes)
            dawn_work = delta
        else: #閉局時間が5:00ちょうどの場合
            sheet['V12'].value = time(hour=0, minute=0)  #'0:00'  同じ場合は0時間
            sheet['V15'].value = time(hour=0, minute=0)  
            dawn_work = timedelta()
    tt_dawn_time = dawn_work + mid_work
    total_minutes = tt_dawn_time.total_seconds() / 60
    sheet['V13'].value = minutes_to_hours_and_minutes(total_minutes)
    tt_usetime_sec = tt_usetime.total_seconds() / 60
    sheet['V19'].value = minutes_to_hours_and_minutes(tt_usetime_sec) #普通の時間外合計時間
    book.save(file_path)

def calculate_work_hours(start_time_str, end_time_str):
    """
    作業開始時間と終了時間を受け取り、総作業時間を計算する。
    作業時間が翌日にまたがる場合や終了時間が24:00から24:59の場合は、適切に処理する。

    :param start_time_str: 作業開始時間の文字列 ('HH:MM'形式)
    :param end_time_str: 作業終了時間の文字列 ('HH:MM'形式)
    :return: 総作業時間を時間と分で返す
    """
    format_str = '%H:%M'  # 時間のフォーマット
    start_time = datetime.strptime(start_time_str, format_str)
    
    # 終了時間が24:00から24:59の場合、それを翌日の時間として扱う
    if end_time_str.startswith('00:'):
        # "24:XX"を"00:XX"として扱い、さらに1日を加算
        adjusted_end_time_str = '00:' + end_time_str.split(':')[1]
        end_time = datetime.strptime(adjusted_end_time_str, format_str) + timedelta(days=1)
    else:
        end_time = datetime.strptime(end_time_str, format_str)

    # 作業終了時間が開始時間より小さい（または等しい）場合、翌日として扱う
    if end_time <= start_time:
        end_time += timedelta(days=1)

    # 総作業時間を計算
    total_work_time = end_time - start_time

    return total_work_time

def format_time_to_str(end):
    # endがdatetime.timeオブジェクトの場合
    if isinstance(end, time):
        return end.strftime('%H:%M')
    # endが文字列の場合
    elif isinstance(end, str):
        try:
            # 文字列が"%H:%M:%S"フォーマットの場合、"%H:%M"にフォーマット
            return datetime.strptime(end, '%H:%M:%S').strftime('%H:%M')
        except ValueError:
            # 文字列が"%H:%M"フォーマットではない場合、またはそれ以外のフォーマットの場合
            # 直接"%H:%M"フォーマットのチェックは行わず、元の文字列を返すか、
            # 追加のフォーマットチェックを行う
            return end
    else:
        # endが文字列でもdatetime.timeオブジェクトでもない場合
        raise TypeError("Unsupported type for 'end'")

def get_datetime_object(time_str):
    """文字列からdatetime.datetimeオブジェクトを生成する"""
    return datetime.strptime(time_str, '%H:%M')

logging.basicConfig(filename='app.log', level=logging.INFO, 
                    format='%(asctime)s %(levelname)s:%(message)s')
# 実際にファイルを更新する場合はこの関数を呼び出します（ファイルパスを適切に設定してください）
# update_overtime_in_excel('your_excel_file.xlsx')
